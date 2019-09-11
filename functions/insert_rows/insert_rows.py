import urllib.parse
import boto3

# For xls
import xlrd # This is much faster than pyexcel and pandas for reading sheet names
            # without having to load the entire document.
            # see: https://stackoverflow.com/questions/12250024/how-to-obtain-sheet-names-from-xls-files-without-loading-the-whole-file

# For xlsx
import openpyxl as opx

import os
import io
from db_helper import *
import queries as q
import datetime
from cash_levels import insert_cl

from sheet_helper import *

file_type = None
FILE_XLS = 0
FILE_XLSX = 1

s3 = boto3.client('s3')

row_number = None # For error handling
sheet_name = None # Set to global for error handling
file_name = None # For error handling

cursor = None
def connect_db(event):
    global cursor
    try:

        cnx = make_connection(event['DBConnection']['endpoint'],
                              event['DBConnection']['port'],
                              event['DBConnection']['dbuser'],
                              event['DBConnection']['dbpassword'],
                              event['DBConnection']['database'])
        cursor=cnx.cursor()
        logger.info('Connected!')
    except:
        return log_err("ERROR: Cannot connect to database from handler.\n{}".format(
            traceback.format_exc()))


def read_object(bucket, key):
    try:
        response = s3.get_object(Bucket=bucket, Key=key)
        return response['Body'].read()
    except Exception as e:
        logger.error(e)
        log_err('Error getting object {} from bucket {}. Make sure they exist and your bucket is in the same region as this function.'.format(key, bucket))
        raise e


def convert_date(value):
    global file_type, row_number, sheet_name, file_name
    try:
        if file_type == FILE_XLS:
            return "'{}'".format(xlrd.xldate.xldate_as_datetime(value, 0).strftime('%Y-%m-%d'))
        elif file_type == FILE_XLSX:
            return "'{}'".format(value.strftime('%Y-%m-%d'))
    except AttributeError as error:
        raise ValueError("Incorrect value for date: {}. See row {} of sheet '{}' in file '{}'".format(value, row_number, sheet_name, file_name))


def convert_int(value):
    global row_number, sheet_name, file_name
    try:
        if isnumeric(value):
            return int(value)
        else:
            return 0
    except ValueError as error:
        raise ValueError("Incorrect value for integer: {}. See row {} of sheet '{}' in file '{}'".format(value, row_number, sheet_name, file_name))


def convert_string(value):
    return "{}".format(str(value))
    

def convert_type(value):
    # Change types like 'EQUITY' to 0
    if value == 'EQUITY':
        return 0
    else:
        return -1


def isnumeric(value):
    try:
        int(value)
        return True
    except ValueError:
        return False


def verify_row(row):
    global file_type
    if file_type == FILE_XLS:
        if row[0].value == '':
            return False
        else:
            return True
    elif file_type == FILE_XLSX:
        if row[0] is None:
            return False
        else:
            return True


new_field_switch = {
    0: convert_date,
    1: convert_string,
    2: convert_type
}


def write_os(fieldmap, row, old_fields):
    global new_field_switch, file_type
    # XLS
    # if row[1].value == 'IATG':
    #     print("IATG!!!")
    #     print(row[old_fields.index('To No of Shares')])
    #     print(row[old_fields.index('To No of Shares')].ctype == xlrd.XL_CELL_ERROR)
    # if row[1].value == 'IATA':
    #     print("IATA!!!")
    #     print(row[old_fields.index('To No of Shares')])
    #
    # XLSX
    # if row[1] == 'IATG':
    #     print("IATG!!!")
    #     print(row[old_fields.index('To No of Shares')])

    newrow_list = []
    for new_field_idx, new_field in enumerate(fieldmap):
        if new_field in old_fields:
            old_field_idx = old_fields.index(new_field)
            fn = new_field_switch.get(new_field_idx, convert_int)
            
            if file_type == FILE_XLS:
                if row[old_field_idx].ctype == xlrd.XL_CELL_ERROR:
                    newrow_list.append(fn(''))
                else:
                    newrow_list.append(fn(row[old_field_idx].value))
            elif file_type == FILE_XLSX:
                newrow_list.append(fn(row[old_field_idx]))
        else:
            newrow_list.append(None)
    # XLS
    # if row[1].value == 'IATG':
    #     print("IATG!!!")
    #     print(newrow_list[23])
    #
    # XLSX
    # if row[1] == 'IATG':
    #     print("IATG!!!")
    #     print(newrow_list[23])

    return newrow_list


def write_lmfo(rowdate, total_id, fieldmap, row):
    global file_type
    if file_type == FILE_XLS:
        return [convert_string(rowdate), convert_string(total_id), convert_string(row[0].value),
                convert_string(row[1].value), convert_int(row[2].value)]
    elif file_type == FILE_XLSX:
        return [convert_string(rowdate), convert_string(total_id), convert_string(row[0]),
                convert_string(row[1]), convert_int(row[2])]


def insert_lmfo(sheet, total_rows, sheet_name):
    global row_number
    insert_str = ""
    fieldmap = []
    inputs = []

    # Get date from sheet name
    dates = sheet_name.split()
    value = "20{}-{}-01".format(dates[1], dates[0])
    rowdate = datetime.datetime.strptime(value, '%Y-%b-%d').strftime('%Y-%m-%d')
    
    # Insert totals
    total_aum = convert_int(get_cell(sheet, 2, 11))
    insert_total_str = q.insert_lmfot_row.format("(%s, %s)")
    total_inputs = (convert_string(rowdate), total_aum)
    result = execute_query(cursor, insert_total_str, inputs=total_inputs)
    if result:
        total_id = result[0][0]
        # print(total_id)
    
        if file_type == FILE_XLS:
            for row_idx in range(1, total_rows):
                row_number = row_idx+2
                row = get_row(sheet, row_idx)
                if verify_row(row):
                    newrow_list = write_lmfo(rowdate, total_id, fieldmap, row)
                    inputs.extend(newrow_list)
                    newrow_str = "(" + ", ".join(['%s']*len(newrow_list)) + ")"
                    insert_str += newrow_str + ", \n"
        elif file_type == FILE_XLSX:
            row_idx = 0
            for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, values_only=True):
                row_number = row_idx+2
                if verify_row(row):
                    newrow_list = write_lmfo(rowdate, total_id, fieldmap, row)
                    inputs.extend(newrow_list)
                    newrow_str = "(" + ", ".join(['%s']*len(newrow_list)) + ")"
                    insert_str += newrow_str + ", \n"
                row_idx += 1
        insert_str = insert_str[:-3]
        insert_str = q.insert_lmfo_rows.format(insert_str)
    return (insert_str, inputs)


def insert_so(sheet, total_rows, old_fields):
    global file_type, row_number
    insert_str = ""
    inputs = []
    fieldmap = q.os_map
    if file_type == FILE_XLS:
        for row_idx in range(1, total_rows):
            row_number = row_idx+2
            row = get_row(sheet, row_idx)
            
            if verify_row(row):
                newrow_list = write_os(fieldmap, row, old_fields)
                inputs.extend(newrow_list)
                newrow_str = "(" + ", ".join(['%s']*len(newrow_list)) + ")"
                insert_str += newrow_str + ", \n"
    elif file_type ==FILE_XLSX:
        row_idx = 0
        for row in sheet.iter_rows(min_row=2, min_col=1, max_col=sheet.max_column, values_only=True):
            row_number = row_idx+2
            if verify_row(row):
                newrow_list = write_os(fieldmap, row, old_fields)
                inputs.extend(newrow_list)
                newrow_str = "(" + ", ".join(['%s']*len(newrow_list)) + ")"
                insert_str += newrow_str + ", \n"
            row_idx += 1
    insert_str = insert_str[:-3]
    insert_str = q.insert_os_rows.format(insert_str)
    return (insert_str, inputs)


# -----------------------
# Sheet-related functions
# -----------------------

def load_sheet(bucket, key, sheet_name):
    global file_type
    if file_type == FILE_XLS:
        wb = xlrd.open_workbook(file_contents=read_object(bucket, key), on_demand=True)
        sheet = wb.sheet_by_name(sheet_name)
    elif file_type == FILE_XLSX:
        wb = opx.load_workbook(io.BytesIO(read_object(bucket, key)), read_only=True, data_only=True)
        sheet = wb[sheet_name]
    return sheet, wb


def get_total_rows(sheet):
    global file_type
    if file_type == FILE_XLS:
        return sheet.nrows
    elif file_type == FILE_XLSX:
        return sheet.max_row
        
        
def get_row(sheet, idx):
    global file_type
    if file_type == FILE_XLS:
        return sheet.row(idx)
    elif file_type == FILE_XLSX:
        # iter_rows is 1-based
        idx += 1
        
        rows = list(sheet.iter_rows(min_row=idx, max_row=idx+1, min_col=0, max_col=sheet.max_column, values_only=False))
        return rows[0]


def get_cell(sheet, row, col):
    # First row and first cell == 1
    global file_type
    if file_type == FILE_XLS:
        return sheet.cell(row-1, col).value
    elif file_type == FILE_XLSX:
        return sheet.cell(row=row, column=col+1).value


def get_num_rows(table_name):
    query = "SELECT COUNT(*) FROM {}".format(table_name)
    result = execute_query(cursor, query)
    return result[0][0]

def lambda_handler(event, context):
    global file_type, sheet_name, file_name
    result = {}
    if 'Records' in event and 'Sheets' in event:
        connect_db(event)
                
        bucket = event['Records'][0]['s3']['bucket']['name']
        key = urllib.parse.unquote_plus(
            event['Records'][0]['s3']['object']['key'],
            encoding='utf-8')
        file_name = event['Records'][0]['s3']['object']['key']
        sheet_name = event['Sheets'][0]

        if key[-3:].lower() == 'xls':
            file_type = FILE_XLS
        elif key[-4:].lower() == 'xlsx':
            file_type = FILE_XLSX
        else:
            log_err('Please upload XLS or XLSX file'.format(key, bucket))

        logger.info("Loading the sheet...")
        sheet, wb = load_sheet(bucket, key, sheet_name)
        sheet_type = check_sheet_type(key)
        logger.info("Sheet loaded! (type: {})".format(sheet_type))
        
        if sheet_type == CASH_LEVELS:
            insert_str, insert_vars = insert_cl(sheet, file_type)
            result['numrows_before_input'] = get_num_rows('catalyst.cash_levels_master')
        else:
            total_rows = get_total_rows(sheet)
    
            # Get fields
            old_fields = list(map(lambda c: c.value, get_row(sheet, 0)))
    
            if sheet_type == LOCAL_MF_OWNERSHIPS:
                insert_str, insert_vars = insert_lmfo(sheet, total_rows, sheet_name)
                result['numrows_before_input'] = get_num_rows('catalyst.local_mf_ownerships_master')
                result['numrows_totals_before_input'] = get_num_rows('catalyst.local_mf_ownership_totals_master')
            elif sheet_type == OWNERSHIP_STRUCTURES:
                insert_str, insert_vars = insert_so(sheet, total_rows, old_fields)
                result['numrows_before_input'] = get_num_rows('catalyst.ownership_structures_master')

        if insert_str != '':
            # logger.info(insert_str)
            logger.info("About to insert...")
            execute_query(cursor, insert_str, inputs=insert_vars, get_result=False)
            logger.info("Success!")
        else:
            logger.info("No valid row, so we don't insert anything.")

        if sheet_type == LOCAL_MF_OWNERSHIPS:
            insert_str, insert_vars = insert_lmfo(sheet, total_rows, sheet_name)
            result['numrows_after_input'] = get_num_rows('catalyst.local_mf_ownerships_master')
            result['numrows_totals_after_input'] = get_num_rows('catalyst.local_mf_ownership_totals_master')
        elif sheet_type == OWNERSHIP_STRUCTURES:
            result['numrows_after_input'] = get_num_rows('catalyst.ownership_structures_master')
        elif sheet_type == CASH_LEVELS:
            result['numrows_after_input'] = get_num_rows('catalyst.cash_levels_master')

        if file_type == FILE_XLS:
            wb.release_resources()
        elif file_type == FILE_XLSX:
            wb.close()

        result['success'] = True
        cursor.close()
        return result
    else:
        result['success'] = False
        return result
